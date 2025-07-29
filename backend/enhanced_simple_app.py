#!/usr/bin/env python3
"""
Al Fozan Insights Platform - Enhanced Backend Server
Real Estate Business Intelligence API with comprehensive features
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime, timedelta
import json
import os
import io
import csv
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app, origins=["http://localhost:3000", "https://alfozan-insights.vercel.app"])

# Create reports directory
REPORTS_DIR = "reports"
if not os.path.exists(REPORTS_DIR):
    os.makedirs(REPORTS_DIR)

# Mock data storage
projects_data = [
    {
        "id": 1,
        "name": "Al Nakheel Towers",
        "location": "Riyadh",
        "type": "Residential",
        "status": "In Progress",
        "completion": 75,
        "budget": 250000000,
        "start_date": "2023-01-15",
        "expected_completion": "2024-12-31"
    },
    {
        "id": 2,
        "name": "Desert Oasis Mall",
        "location": "Jeddah",
        "type": "Commercial",
        "status": "Planning",
        "completion": 25,
        "budget": 180000000,
        "start_date": "2023-06-01",
        "expected_completion": "2025-08-15"
    },
    {
        "id": 3,
        "name": "Coastal Residences",
        "location": "Dammam",
        "type": "Residential",
        "status": "Completed",
        "completion": 100,
        "budget": 120000000,
        "start_date": "2022-03-10",
        "expected_completion": "2023-11-30"
    }
]

competitors_data = [
    {
        "id": 1,
        "name": "Saudi Real Estate Co.",
        "market_share": 15.2,
        "digital_presence": 85,
        "recent_projects": 12,
        "strength": "Market Leadership"
    },
    {
        "id": 2,
        "name": "Gulf Development Group",
        "market_share": 12.8,
        "digital_presence": 78,
        "recent_projects": 8,
        "strength": "Innovation Focus"
    },
    {
        "id": 3,
        "name": "Arabian Construction Ltd",
        "market_share": 10.5,
        "digital_presence": 72,
        "recent_projects": 15,
        "strength": "Cost Efficiency"
    }
]

def calculate_analytics():
    """Calculate comprehensive analytics from project data"""
    total_projects = len(projects_data)
    total_revenue = sum(project["budget"] for project in projects_data) / 1000000  # Convert to millions
    total_units = sum(project["units"] for project in projects_data)
    total_units_sold = sum(project["units_sold"] for project in projects_data)
    sales_rate = (total_units_sold / total_units * 100) if total_units > 0 else 0
    
    # Calculate additional metrics
    avg_project_value = total_revenue / total_projects if total_projects > 0 else 0
    completion_rate = len([p for p in projects_data if p["status"] == "Completed"]) / total_projects * 100
    
    return {
        "kpis": {
            "total_projects": total_projects,
            "total_revenue": total_revenue,
            "total_units": total_units,
            "total_units_sold": total_units_sold,
            "sales_rate": round(sales_rate, 1),
            "avg_project_value": round(avg_project_value, 1),
            "completion_rate": round(completion_rate, 1)
        },
        "regional_data": [
            {"region": "Riyadh", "projects": 1, "revenue": 450, "units_sold": 127},
            {"region": "Jeddah", "projects": 1, "revenue": 680, "units_sold": 8},
            {"region": "Dammam", "projects": 1, "revenue": 320, "units_sold": 23},
            {"region": "Mecca", "projects": 1, "revenue": 280, "units_sold": 34},
            {"region": "Al-Khobar", "projects": 1, "revenue": 195, "units_sold": 28}
        ],
        "project_types": [
            {"type": "Residential", "count": 2, "revenue": 645, "percentage": 35.8},
            {"type": "Commercial", "count": 2, "revenue": 960, "percentage": 53.3},
            {"type": "Industrial", "count": 1, "revenue": 320, "percentage": 17.8}
        ]
    }

def generate_pdf_report(report_type="comprehensive", date_range="month"):
    """Generate comprehensive PDF report"""
    filename = f"alfozan_report_{report_type}_{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    # Create PDF document
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.darkblue
    )
    
    # Title page
    story.append(Paragraph("Al Fozan Holding", title_style))
    story.append(Paragraph("Real Estate Insights Platform", styles['Heading2']))
    story.append(Spacer(1, 0.5*inch))
    story.append(Paragraph(f"Business Intelligence Report - {report_type.title()}", heading_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
    story.append(Paragraph(f"Report Period: {date_range.title()}", styles['Normal']))
    story.append(PageBreak())
    
    # Executive Summary
    story.append(Paragraph("Executive Summary", heading_style))
    analytics = calculate_analytics()
    kpis = analytics["kpis"]
    
    summary_data = [
        ["Metric", "Value", "Performance"],
        ["Total Projects", str(kpis["total_projects"]), "Active Portfolio"],
        ["Total Revenue", f"{kpis['total_revenue']:.1f}M SAR", "Strong Performance"],
        ["Units Sold", f"{kpis['total_units_sold']}/{kpis['total_units']}", f"{kpis['sales_rate']:.1f}% Sales Rate"],
        ["Completion Rate", f"{kpis['completion_rate']:.1f}%", "On Track"],
        ["Avg Project Value", f"{kpis['avg_project_value']:.1f}M SAR", "Premium Segment"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Project Portfolio Analysis
    story.append(Paragraph("Project Portfolio Analysis", heading_style))
    
    project_data = [["Project Name", "Type", "Status", "Progress", "Budget (M SAR)"]]
    for project in projects_data:
        project_data.append([
            project["name"],
            project["type"],
            project["status"],
            f"{project['completion']}%",
            f"{project['budget']/1000000:.1f}"
        ])
    
    project_table = Table(project_data, colWidths=[2.5*inch, 1*inch, 1*inch, 0.8*inch, 1*inch])
    project_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9)
    ]))
    
    story.append(project_table)
    story.append(PageBreak())
    
    # Market Analysis
    story.append(Paragraph("Competitive Market Analysis", heading_style))
    
    competitor_data = [["Competitor", "Market Share", "Digital Score", "Trend"]]
    for competitor in competitors_data:
        competitor_data.append([
            competitor["name"],
            f"{competitor['market_share']}%",
            f"{competitor['digital_presence']}/100",
            f"{competitor['strength']}"
        ])
    
    competitor_table = Table(competitor_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1.5*inch])
    competitor_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9)
    ]))
    
    story.append(competitor_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Strategic Recommendations
    story.append(Paragraph("Strategic Recommendations", heading_style))
    recommendations = [
        "• Focus on completing high-progress projects to improve cash flow",
        "• Accelerate sales efforts for projects with low sales rates",
        "• Consider market expansion in underserved regions",
        "• Enhance digital presence to compete with market leaders",
        "• Diversify project portfolio to reduce market risk"
    ]
    
    for rec in recommendations:
        story.append(Paragraph(rec, styles['Normal']))
        story.append(Spacer(1, 6))
    
    # Build PDF
    doc.build(story)
    
    return filename

def generate_excel_report(report_type="comprehensive", date_range="month"):
    """Generate comprehensive Excel report"""
    filename = f"alfozan_data_{report_type}_{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Summary Sheet
    summary_ws = wb.create_sheet("Executive Summary")
    analytics = calculate_analytics()
    kpis = analytics["kpis"]
    
    # Headers
    summary_ws['A1'] = "Al Fozan Holding - Executive Summary"
    summary_ws['A1'].font = Font(size=16, bold=True)
    summary_ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # KPIs
    summary_ws['A4'] = "Key Performance Indicators"
    summary_ws['A4'].font = Font(size=14, bold=True)
    
    kpi_data = [
        ["Metric", "Value"],
        ["Total Projects", kpis["total_projects"]],
        ["Total Revenue (M SAR)", f"{kpis['total_revenue']:.1f}"],
        ["Total Units", kpis["total_units"]],
        ["Units Sold", kpis["total_units_sold"]],
        ["Sales Rate (%)", f"{kpis['sales_rate']:.1f}"],
        ["Average Project Value (M SAR)", f"{kpis['avg_project_value']:.1f}"],
        ["Completion Rate (%)", f"{kpis['completion_rate']:.1f}"]
    ]
    
    for row_idx, row_data in enumerate(kpi_data, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 5:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
    
    # Projects Sheet
    projects_ws = wb.create_sheet("Projects")
    project_headers = ["ID", "Name", "Type", "Status", "Location", "Budget", "Completion", "Start Date", "Expected Completion"]
    
    for col_idx, header in enumerate(project_headers, start=1):
        cell = projects_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_idx, project in enumerate(projects_data, start=2):
        projects_ws.cell(row=row_idx, column=1, value=project["id"])
        projects_ws.cell(row=row_idx, column=2, value=project["name"])
        projects_ws.cell(row=row_idx, column=3, value=project["type"])
        projects_ws.cell(row=row_idx, column=4, value=project["status"])
        projects_ws.cell(row=row_idx, column=5, value=project["location"])
        projects_ws.cell(row=row_idx, column=6, value=project["budget"])
        projects_ws.cell(row=row_idx, column=7, value=project["completion"])
        projects_ws.cell(row=row_idx, column=8, value=project["start_date"])
        projects_ws.cell(row=row_idx, column=9, value=project["expected_completion"])
    
    # Competitors Sheet
    competitors_ws = wb.create_sheet("Competitors")
    competitor_headers = ["ID", "Name", "Market Share", "Digital Presence", "Recent Projects", "Strength"]
    
    for col_idx, header in enumerate(competitor_headers, start=1):
        cell = competitors_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_idx, competitor in enumerate(competitors_data, start=2):
        competitors_ws.cell(row=row_idx, column=1, value=competitor["id"])
        competitors_ws.cell(row=row_idx, column=2, value=competitor["name"])
        competitors_ws.cell(row=row_idx, column=3, value=competitor["market_share"])
        competitors_ws.cell(row=row_idx, column=4, value=competitor["digital_presence"])
        competitors_ws.cell(row=row_idx, column=5, value=competitor["recent_projects"])
        competitors_ws.cell(row=row_idx, column=6, value=competitor["strength"])
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(filepath)
    
    return filename

def generate_csv_report(report_type="comprehensive", date_range="month"):
    """Generate CSV data export"""
    filename = f"alfozan_data_{report_type}_{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
        if report_type == "projects" or report_type == "comprehensive":
            writer = csv.writer(csvfile)
            
            # Write projects data
            writer.writerow(["# PROJECTS DATA"])
            writer.writerow(["ID", "Name", "Location", "Type", "Status", "Completion", "Budget", "Start_Date", "Expected_Completion"])
            
            for project in projects_data:
                writer.writerow([
                    project["id"],
                    project["name"],
                    project["location"],
                    project["type"],
                    project["status"],
                    project["completion"],
                    project["budget"],
                    project["start_date"],
                    project["expected_completion"]
                ])
            
            writer.writerow([])  # Empty row
            
        if report_type == "competitors" or report_type == "comprehensive":
            writer.writerow(["# COMPETITORS DATA"])
            writer.writerow(["ID", "Name", "Market_Share", "Digital_Presence", "Recent_Projects", "Strength"])
            
            for competitor in competitors_data:
                writer.writerow([
                    competitor["id"],
                    competitor["name"],
                    competitor["market_share"],
                    competitor["digital_presence"],
                    competitor["recent_projects"],
                    competitor["strength"]
                ])
    
    return filename

# API Routes

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0"
    })

@app.route('/api/auth/login', methods=['POST'])
def login():
    """User authentication endpoint"""
    data = request.get_json()
    username = data.get('username', '')
    password = data.get('password', '')
    
    # Demo credentials
    demo_users = {
        'admin': {'password': 'admin123', 'role': 'admin'},
        'manager': {'password': 'manager123', 'role': 'manager'},
        'analyst': {'password': 'analyst123', 'role': 'analyst'}
    }
    
    if username in demo_users and demo_users[username]['password'] == password:
        return jsonify({
            "success": True,
            "user": {
                "username": username,
                "role": demo_users[username]['role'],
                "token": f"demo_token_{username}_{datetime.now().timestamp()}"
            }
        })
    
    return jsonify({"success": False, "message": "Invalid credentials"}), 401

@app.route('/api/projects', methods=['GET'])
def get_projects():
    """Get all projects"""
    try:
        return jsonify({"projects": projects_data})
    except Exception as e:
        logger.error(f"Get projects error: {str(e)}")
        return jsonify({
            "success": False,
            "error": "Failed to fetch projects"
        }), 500

@app.route('/api/projects', methods=['POST'])
def create_project():
    """Create new project"""
    try:
        data = request.get_json()
        
        # Generate new ID
        new_id = len(projects_data) + 1
        
        # Create new project
        new_project = {
            "id": new_id,
            "name": data.get("name", ""),
            "location": data.get("location", ""),
            "type": data.get("type", ""),
            "status": data.get("status", "Planning"),
            "completion": data.get("completion", 0),
            "budget": data.get("budget", 0),
            "start_date": data.get("start_date", datetime.now().strftime('%Y-%m-%d')),
            "expected_completion": data.get("expected_completion", "")
        }
        
        projects_data.append(new_project)
        
        return jsonify({
            "success": True,
            "project": new_project
        }), 201
        
    except Exception as e:
        logger.error(f"Create project error: {str(e)}")
        return jsonify({
            "success": False,
            "error": "Failed to create project"
        }), 500

@app.route('/api/projects/<int:project_id>', methods=['PUT'])
def update_project(project_id):
    """Update existing project"""
    try:
        data = request.get_json()
        
        # Find project
        project = next((p for p in projects_data if p["id"] == project_id), None)
        if not project:
            return jsonify({
                "success": False,
                "error": "Project not found"
            }), 404
        
        # Update project fields
        project.update(data)
        
        return jsonify({
            "success": True,
            "project": project
        })
        
    except Exception as e:
        logger.error(f"Update project error: {str(e)}")
        return jsonify({
            "success": False,
            "error": "Failed to update project"
        }), 500

@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    """Delete project"""
    try:
        # Find and remove project
        global projects_data
        projects_data = [p for p in projects_data if p['id'] != project_id]
        
        return jsonify({
            "success": True
        })
        
    except Exception as e:
        logger.error(f"Delete project error: {str(e)}")
        return jsonify({
            "success": False,
            "error": "Failed to delete project"
        }), 500

@app.route('/api/competitors', methods=['GET'])
def get_competitors():
    """Get all competitors"""
    try:
        return jsonify({"competitors": competitors_data})
    except Exception as e:
        logger.error(f"Get competitors error: {str(e)}")
        return jsonify({
            "success": False,
            "error": "Failed to fetch competitors"
        }), 500

@app.route('/api/competitors', methods=['POST'])
def create_competitor():
    """Create new competitor"""
    try:
        data = request.get_json()
        
        # Generate new ID
        new_id = len(competitors_data) + 1
        
        # Create new competitor
        new_competitor = {
            "id": new_id,
            "name": data.get("name", ""),
            "market_share": data.get("market_share", 0),
            "digital_presence": data.get("digital_presence", 0),
            "recent_projects": data.get("recent_projects", 0),
            "strength": data.get("strength", "")
        }
        
        competitors_data.append(new_competitor)
        
        return jsonify({
            "success": True,
            "competitor": new_competitor
        }), 201
        
    except Exception as e:
        logger.error(f"Create competitor error: {str(e)}")
        return jsonify({
            "success": False,
            "error": "Failed to create competitor"
        }), 500

@app.route('/api/analytics/dashboard', methods=['GET'])
def get_dashboard_analytics():
    """Get dashboard analytics"""
    total_projects = len(projects_data)
    active_projects = len([p for p in projects_data if p['status'] == 'In Progress'])
    total_budget = sum(p['budget'] for p in projects_data)
    avg_completion = sum(p['completion'] for p in projects_data) / len(projects_data) if projects_data else 0
    
    return jsonify({
        "kpis": {
            "total_projects": total_projects,
            "active_projects": active_projects,
            "total_budget": total_budget,
            "avg_completion": round(avg_completion, 1)
        },
        "regional_performance": [
            {"region": "Riyadh", "projects": 8, "value": 450000000},
            {"region": "Jeddah", "projects": 5, "value": 320000000},
            {"region": "Dammam", "projects": 3, "value": 180000000}
        ],
        "monthly_trends": [
            {"month": "Jan", "projects": 2, "revenue": 45000000},
            {"month": "Feb", "projects": 3, "revenue": 52000000},
            {"month": "Mar", "projects": 1, "revenue": 38000000},
            {"month": "Apr", "projects": 4, "revenue": 67000000},
            {"month": "May", "projects": 2, "revenue": 41000000},
            {"month": "Jun", "projects": 3, "revenue": 58000000}
        ]
    })

@app.route('/api/export/projects', methods=['GET'])
def export_projects():
    """Export projects data"""
    format_type = request.args.get('format', 'json')
    
    if format_type == 'csv':
        output = io.StringIO()
        output.write("ID,Name,Location,Type,Status,Completion,Budget\n")
        for project in projects_data:
            output.write(f"{project['id']},{project['name']},{project['location']},{project['type']},{project['status']},{project['completion']},{project['budget']}\n")
        
        return jsonify({
            "success": True,
            "data": output.getvalue(),
            "filename": f"projects_export_{datetime.now().strftime('%Y%m%d')}.csv"
        })
    
    return jsonify({"projects": projects_data})

@app.route('/api/download/<filename>', methods=['GET'])
def download_file(filename):
    """Download generated reports"""
    try:
        filepath = os.path.join(REPORTS_DIR, filename)
        if not os.path.exists(filepath):
            return jsonify({
                "success": False,
                "error": "File not found"
            }), 404
        
        return send_file(filepath, as_attachment=True)
        
    except Exception as e:
        logger.error(f"Download file error: {str(e)}")
        return jsonify({
            "success": False,
            "error": "Failed to download file"
        }), 500

if __name__ == '__main__':
    print("🚀 Starting Al Fozan Insights Platform Backend...")
    print("📊 Dashboard: http://localhost:5000/api/health")
    print("🔐 Demo Credentials:")
    print("   Admin: admin/admin123")
    print("   Manager: manager/manager123") 
    print("   Analyst: analyst/analyst123")
    app.run(debug=True, host='0.0.0.0', port=5000)
